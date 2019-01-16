#Gestion de archivos excel con la liberia openpyxl
#se instala con: pip install openpyxl
#python
# import openpyxl 



def p(mensaje):
	print(mensaje+"")

p('**** XLS2SQL : GENERADOR DE SENTENCIAS SQL DESDE EXCEL   ****')


import openpyxl
import os 
import sys 

#MMETODO DE INVOCACION 
# python xlsx2sql.py core_20181201.xlsx 20181201



# P A R A M E T R O S
#cantidad de parametros

#print("Cantidad"+str(len(sys.argv)))

#Nombre de Archivos de entrada (xlsx)
nombre_archivo_excel=sys.argv[1] #"core_20181201.xlsx" # parametro 1

#fecha del proceso
fecha_archivo=sys.argv[2] #"20181201" #parametro 2

#nombre de archivo salida (sql)
if len(sys.argv)==4:
	nombre_archivo_sql=sys.argv[3]
else:
	nombre_archivo_sql=nombre_archivo_excel.replace("xlsx","sql")


# Log
#p('Archivo ENTRADA (EXCEL): '+nombre_archivo_excel) #+ "  / HOJA  :" + nombre_hoja)
#p('Fecha de Proceso       : '+fecha_archivo)
#p('Archivo SALIDA  (SQL)  : '+nombre_archivo_sql)

p("Conversion XLSX A SQL:"+ nombre_archivo_excel + " --> " + nombre_archivo_sql)

#Apertura de Archivo Entrada
wb = openpyxl.load_workbook(nombre_archivo_excel, data_only=True) #lee excel con la formula aplicada

#Carga de Hoja de Procesamiento (la primera hoja)
nombre_hoja=wb.sheetnames[0]
hoja = wb[nombre_hoja]

#Apertura de Archivo Salida
archivo_sql=open(nombre_archivo_sql,mode="w")


# Cabecera de archivo de Salida
archivo_sql.write('-- Archivo ENTRADA (EXCEL): '+nombre_archivo_excel+ "  / HOJA  :" + nombre_hoja + "\n")
archivo_sql.write('-- Archivo SALIDA  (SQL)  : '+nombre_archivo_sql + "\n\n")
archivo_sql.write("--Configuracion de Sesion"+"\n")
archivo_sql.write("SET FEEDBACK OFF" + "\n\n")
archivo_sql.write("ALTER SESSION SET NLS_DATE_FORMAT='YYYY/MM/DD HH24:MI:SS' ;" + "\n\n")
archivo_sql.write("--Reseteo de Datos"+"\n")
#archivo_sql.write("DELETE ctrlm_stats WHERE fecha_excel =" + fecha_archivo+ ";\n\n")
#rangos de hoja
fila_min = hoja.min_row 
fila_max = hoja.max_row
col_min = hoja.min_column
col_max = hoja.max_column
#
#print(f"Fila Minima --> {fila_min}")
#print(f"Fila Maxima --> {fila_max}")
#print(f"Columna Minima --> {col_min}")
#print(f"Columna Maxima --> {col_max}")

#Definición de Estructura de Estadistica de Malla
# Columna A 0   Malla	
# Columna B 1   Grupo	
# Columna C 2   Proceso
# Columna D 3   Job
# Columna E 4   Nodo
# Columna F 5   CTRLM Inicio
# Columna G 6   CTRLM Demora
# Columna H 7   Inicio
# Columna I 8   Fin
# Columna J 9   Demora

# insert into ctrlm_stats 
# (fecha_excel,malla,grupo,proceso,job,nodo,inicio_n,duracion_n,inicio_h,fin_h,duracion_t) values (

# =INSERT &$K$1 &",'"&A2&"','"&B2&"','"&C2&"','"&D2&"','"&E2&"',"&F2&","&G2&",'"&TEXTO(H2,"YYYY/MM/DD HH:MM:SS")&"','"&I2&"','"&TEXTO(J2,"HH:MM:SS")&"');"

# -- constantes iniciales

p(' ... Inicio Proceso Hoja ' + nombre_hoja)

insert="insert into ctrlm_stats (fecha_excel,malla,grupo,proceso,job,nodo,fin_n,duracion_n,inicio_h,fin_h,duracion_t) values "
values=""

#columnas=range(1,col_max+1)

i=2
#top=fila_max/100
#plantilla='Los valores son: {0}    {1}   {2}'
#print(hoja['A2'].value)
#while (hoja['A'+str(i)].value != ""):
while i<fila_max:
	j=str(i)
	malla=hoja['A'+j].value
	#ignorar los espacios en blanco 
	if (malla is not None):
		#print("linea "+j)
		grupo=hoja['B'+j].value
		proceso=hoja['C'+j].value 
		job=hoja['D'+j].value 
		nodo=hoja['E'+j].value 
		#Valores Control M 
		fin_n=hoja['F'+j].value 
		duracion_n=hoja['G'+j].value 
		#Valores Date
		inicio_f=hoja['H'+j].value 
		fin_t=hoja['I'+j].value
		#Duracion Texto 
		duracion_f=hoja['J'+j].value 
		#conversion a texto
		inicio_t=f'{inicio_f:%Y/%m/%d %H:%M:%S.%f}'
		duracion_t=f'{duracion_f:%H:%M:%S}'
		#print("Inicio f:"+str(inicio_f))  
		values="("+fecha_archivo+",'"+malla+"','"+grupo+"','"+proceso+"','"+job+"','"
		values=values+nodo+"',"+str(fin_n)+","+str(duracion_n)+","
		#esta en formato decimal por lo que primero se pasa a timestamp y luego a date 
		values=values+"CAST(TO_TIMESTAMP('"+inicio_t+"', 'YYYY-MM-DD HH24:MI:SS,FF9') AS DATE),'"
		#values=value+inicio_t+"','"
		values=values+fin_t+"','"+duracion_t+"');"
		#print(plantilla.format(malla,grupo,shell))
		#print(values)
		archivo_sql.write(insert + values +"\n")
	i=i+1

wb.close()
archivo_sql.close()

p(' ... Fin Proceso Hoja ' + nombre_hoja)

exit() 


# p("carga titulos de la hoja")

# titulos=[0,]
# for i in columnas:
# 	titulos.insert(i,hoja.cell(row=1,column=i).value)
# #	print(hoja.cell(row=1,column=i).value)
# print(titulos)

# p("recorrido de hoja e impresion de pantalla I")
# filas=range(2,fila_max+1)
# for i in filas:
# 	for j in columnas:
# 		print(f" {titulos[j]} --> {hoja.cell(row=i,column=j).value}")
# #	print(hoja.cell(row=1,column=i).value)
# 	print("")


# p("recorrido de hoja e impresion en pantalla II")


