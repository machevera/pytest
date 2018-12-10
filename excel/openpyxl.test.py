#Gestion de archivos excel con la liberia openpyxl
#se instala con: pip install openpyxl
#python
# import openpyxl 
# print(openpyxl.__version__)


def p(mensaje):
	print("\n--"+mensaje+"--")

p('** GESTION DE ARCHIVOS EXCEL con openpyxl **')

p('version de openpyxl')
import openpyxl
import os 

print (openpyxl.__version__)

p('abriendo archivo EXCEL')
wb = openpyxl.load_workbook("prueba1.xlsx")
p('tipo de archivo del handler de excel')
print(type(wb))


p('hojas del archivo')
# wb.get_sheet_names() deprecado
hojas=wb.sheetnames
print(hojas)
print(type(hojas))

p('carga de la primera hoja')
hoja = wb["Hoja1"]

p('rango con datos de la primera hoja')
fila_min = hoja.min_row 
fila_max = hoja.max_row
col_min = hoja.min_column
col_max = hoja.max_column
print(f"Fila Minima --> {fila_min}")
print(f"Fila Maxima --> {fila_max}")
print(f"Columna Minima --> {col_min}")
print(f"Columna Maxima --> {col_max}")

p("carga titulos de la hoja")
columnas=range(1,col_max+1)
titulos=[0,]
for i in columnas:
	titulos.insert(i,hoja.cell(row=1,column=i).value)
#	print(hoja.cell(row=1,column=i).value)
print(titulos)

p("recorrido de hoja e impresion de pantalla I")
filas=range(2,fila_max+1)
for i in filas:
	for j in columnas:
		print(f" {titulos[j]} --> {hoja.cell(row=i,column=j).value}")
#	print(hoja.cell(row=1,column=i).value)
	print("")


p("recorrido de hoja e impresion en pantalla II")
i=2
plantilla='Los valores son: {0}    {1}   {2}'
#print(hoja['A2'].value)
#while (hoja['A'+str(i)].value != ""):
while i<fila_max:
	j=str(i)
	print(plantilla.format(hoja['A'+j].value,hoja['B'+j].value,hoja['C'+j].value))
	i=i+1

wb.close()